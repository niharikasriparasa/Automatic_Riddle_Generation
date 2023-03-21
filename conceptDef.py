import wikipedia
import pandas as pd
import xlsxwriter
import openpyxl
import numpy as np
#from pyopenie import OpenIE5
#extractor = OpenIE5('http://localhost:8000')

#read excel sheet
#df=pd.read_excel(r'C:\Users\Niharika Sri Parasa\PycharmProjects\RiddleGen_mod1\topicsList.xlsx',sheet_name='topics_Def')
wb=openpyxl.load_workbook('topicsList.xlsx')
#df=df.astype(str)
wbsheet=wb.get_sheet_by_name('Physics')
rows=wbsheet.max_row
print(rows)
columns=wbsheet.max_column
print(columns)
for i in range(1,rows+1):
    topic = wbsheet.cell(row=i, column=1).value
    if wbsheet.cell(row=i,column=2).value is None :#change column number as accordingly
        try:
            #sen = wikipedia.summary(topic, sentences=1)
            #sen=wikipedia.summary(topic,sentences=3)
            sen = wikipedia.summary(topic)
            #wbsheet.cell(row=i,column=2).value=sen
            #wbsheet.cell(row=i, column=3).value = sen
            wbsheet.cell(row=i, column=2).value = sen
            print(topic,sen)
        except(ConnectionAbortedError):
            print("connection aborted error")
        except(ConnectionRefusedError):
            print("Connection refused error")
        except(ConnectionResetError):
            print("Connection reset error")
        except(wikipedia.exceptions.PageError):
            print("page error")
        except(ConnectionError):
            print("connection error")
        except(wikipedia.exceptions.DisambiguationError):
            print("Disambiguation error")
        except wikipedia.requests.exceptions.ConnectionError:
            print('Http error')
        finally:
            wb.save('topicsList.xlsx')
    else:
            print(' no def ')
"""
for index, row in df.iterrows():
    topic = row['Topics']
    print(topic)
    if (len(row['1-sentence']) == 3) :

        try:
            text = wikipedia.summary(topic, sentences=1)
            workheet
        



"""